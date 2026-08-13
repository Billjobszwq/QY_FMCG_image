"""ABOSV3 T3：Import Center（03-DATA-IMPORT-MANUAL-FALLBACK-AND-CONTRACTS）。

- 全模块共用一个导入中心；14 套 CSV/XLSX 模板（带字段说明/样例/幂等键）；
- 状态机 uploaded → parsed → mapped → dry_run_passed → committed
  （失败：parse_failed / validation_failed）；
- 逐行新增/跳过/冲突/错误分类；提交按自然键幂等；
- 原文件 hash、actor、结果与错误报告进 Evidence + 审计；
- 不得直接写 SQLite 绕 Domain Service：所有提交走 IAM/MasterData/
  Survey/FieldOps/Finance 服务。
"""
from __future__ import annotations

import csv
import hashlib
import io
import json
import uuid
from datetime import datetime, timezone
from typing import Any

from openpyxl import Workbook, load_workbook


class ImportError_(Exception):
    """导入中心错误（诚实失败）。"""


class ImportAuthError(ImportError_):
    """OSV5：导入授权失败（fail-closed，API 映射 403）。"""


class AdjudicationError(ImportError_):
    """OSV51 C-3：隔离区裁决状态机错误（稳定错误码，API 映射 409）。"""


# OSV51 C-2：敏感键名（子串匹配，大小写不敏感，fail-closed 宁多勿漏）。
SECRET_KEY_SUBSTRINGS = ("password", "passwd", "token", "api_key",
                         "apikey", "secret", "credential",
                         "private_key")


def _is_secret_key(key: str) -> bool:
    kl = str(key).lower()
    return any(s in kl for s in SECRET_KEY_SUBSTRINGS)


def redact_secrets(obj: Any) -> Any:
    """OSV51 C-2：递归 secret 扫描/脱敏（DTO 与落库 JSON 共用）。

    dict/list 任意深度；敏感键的字符串值替换为 '[REDACTED]'（保留键
    与非敏感字段）。非字符串值不视为秘密载体，但继续递归其结构。
    """
    if isinstance(obj, dict):
        out: dict = {}
        for k, v in obj.items():
            if _is_secret_key(k) and isinstance(v, str) and v \
                    and v != "[REDACTED]":
                out[k] = "[REDACTED]"
            else:
                out[k] = redact_secrets(v)
        return out
    if isinstance(obj, list):
        return [redact_secrets(v) for v in obj]
    return obj


# OSV5（指令 5.2）：模板 → capability scope 矩阵（版本化 permission
# policy，不得只判断是否登录；全局模板不得因无 customer_id 绕过）。
TEMPLATE_SCOPE: dict[str, str] = {
    "customers_v1": "master.manage",
    "projects_v1": "master.manage",
    "skus_v1": "master.manage",
    "stores_addresses_v1": "master.manage",
    "employees_v1": "master.manage",
    "route_constraints_v1": "master.manage",
    "knowledge_documents_v1": "master.manage",
    "users_v1": "iam.manage",
    "roles_permissions_v1": "iam.manage",
    "memberships_v1": "iam.manage",
    "survey_definition_v1": "survey.manage",
    "survey_questions_v1": "survey.manage",
    "survey_logic_v1": "survey.manage",
    "usage_rate_cards_v1": "finance.manage",
}

# 模板 → 行内客户列（用于批次客户作用域推导；无客户列的模板为 None）。
TEMPLATE_CUSTOMER_COL: dict[str, str | None] = {
    "customers_v1": "customer_id",
    "projects_v1": "customer_id",
    "stores_addresses_v1": "customer_id",
    "employees_v1": "customer_id",
    "route_constraints_v1": "customer_id",
    "memberships_v1": "customer_id",
    "knowledge_documents_v1": "customer_id",
    "skus_v1": None, "users_v1": None, "roles_permissions_v1": None,
    "survey_definition_v1": None, "survey_questions_v1": None,
    "survey_logic_v1": None, "usage_rate_cards_v1": None,
}

# fixture 批次提交后作用域继承（自然键 → 目标表）。
_SCOPE_INHERIT_NATURAL: dict[str, tuple[str, str]] = {
    "customers_v1": ("md_customer_v1", "customer_id"),
    "projects_v1": ("md_project_v1", "project_id"),
    "skus_v1": ("md_sku_v1", "sku_id"),
    "users_v1": ("iam_principal_v1", "username"),
}
_SCOPE_INHERIT_RECEIPT: dict[str, tuple[str, str]] = {
    "stores_addresses_v1": ("geo_address_v1", "address_id"),
    "employees_v1": ("geo_employee_v1", "employee_id"),
}


def _new_id(prefix: str) -> str:
    return f"{prefix}-" + uuid.uuid4().hex[:12]


def _now() -> str:
    return datetime.now(timezone.utc).isoformat()


def _col(field: str, label: str, *, type_: str = "str", required: bool,
         enum: tuple = (), example: str = "", doc: str = "") -> dict:
    return {"field": field, "label": label, "type": type_,
            "required": required, "enum": list(enum),
            "example": example, "doc": doc}


# ---------- 14 套模板定义（03 文档 §1） ----------

TEMPLATES: dict[str, dict] = {
    "customers_v1": {
        "name": "客户主数据",
        "idempotency": "customer_id（同一客户重复导入只更新幂等跳过）",
        "columns": [
            _col("customer_id", "客户ID", required=True,
                 example="cust-demo", doc="全局唯一客户标识（幂等键）"),
            _col("name", "客户名称", required=True, example="演示客户"),
            _col("payment_terms", "账期", required=False, example="月结30天"),
            _col("retention_policy", "数据保留策略", required=False,
                 enum=["", "半年", "2年", "7年"], example="2年"),
            _col("tags", "标签", required=False, example="快消;连锁",
                 doc="分号分隔"),
        ],
    },
    "projects_v1": {
        "name": "项目主数据",
        "idempotency": "project_id",
        "columns": [
            _col("project_id", "项目ID", required=True, example="prj-demo-1"),
            _col("customer_id", "客户ID", required=True, example="cust-demo"),
            _col("name", "项目名称", required=True, example="华东货架巡检"),
            _col("start_date", "开始日期", type_="date", required=False,
                 example="2026-08-01"),
            _col("end_date", "结束日期", type_="date", required=False,
                 example="2026-09-30"),
            _col("owner", "负责人", required=False, example="bill"),
            _col("budget", "预算", type_="float", required=False,
                 example="50000"),
        ],
    },
    "skus_v1": {
        "name": "SKU 主数据",
        "idempotency": "sku_id",
        "columns": [
            _col("sku_id", "SKU ID", required=True, example="sku-cola-500"),
            _col("canonical_name", "标准名称", required=True,
                 example="可乐 500ml"),
            _col("brand", "品牌", required=False, example="示例品牌"),
            _col("category", "品类", required=False, example="碳酸饮料"),
            _col("volume", "容量", required=False, example="500ml"),
            _col("package_version", "包装版本", required=False,
                 example="v1", doc="新旧包装用 supersede 链，不覆盖历史"),
            _col("barcode", "条码", required=False, example="6900000000001"),
            _col("aliases", "别名", required=False, example="大可乐;cola500",
                 doc="分号分隔"),
            _col("valid_from", "生效日期", type_="date", required=False,
                 example="2026-01-01"),
            _col("valid_to", "失效日期", type_="date", required=False,
                 example=""),
        ],
    },
    "stores_addresses_v1": {
        "name": "门店与地址",
        "idempotency": "customer_id + store_name",
        "columns": [
            _col("customer_id", "客户ID", required=True, example="cust-demo"),
            _col("store_name", "门店名称", required=True, example="示例便利店"),
            _col("raw_address", "标准地址", required=True,
                 example="上海市浦东新区示例路 100 号"),
            _col("region", "区域", required=False, example="华东"),
            _col("lat", "纬度", type_="float", required=False,
                 example="31.2304", doc="无坐标可留空，稍后用获取坐标"),
            _col("lng", "经度", type_="float", required=False,
                 example="121.4737"),
            _col("coord_system", "坐标系", required=False,
                 enum=["", "wgs84", "gcj02"], example="wgs84"),
            _col("time_window", "营业时间窗", required=False,
                 example="09:00-18:00"),
        ],
    },
    "employees_v1": {
        "name": "外勤员工",
        "idempotency": "customer_id + name",
        "columns": [
            _col("customer_id", "客户ID", required=True, example="cust-demo"),
            _col("name", "员工姓名", required=True, example="张三"),
            _col("skills", "技能", required=False, example="巡检;拍照",
                 doc="分号分隔"),
            _col("vehicle", "交通工具", required=False, example="电动车"),
            _col("shift", "班次", required=False, example="早班"),
            _col("home_base", "起点地址", required=False,
                 example="上海市示例集散点"),
        ],
    },
    "users_v1": {
        "name": "平台用户",
        "idempotency": "username",
        "columns": [
            _col("username", "用户名", required=True, example="alice"),
            _col("display_name", "显示名", required=False, example="Alice"),
            _col("kind", "身份类型", required=False,
                 enum=["", "user", "service_account"], example="user"),
            _col("status", "状态", required=False,
                 enum=["", "active", "disabled"], example="active"),
        ],
        "note": "提交时为每个用户生成一次性初始口令：仅在提交成功的当次"
                "响应中返回一次，系统不落库明文（DB 仅存哈希，回执以"
                "[REDACTED] 存档），用户应尽快修改。",
    },
    "roles_permissions_v1": {
        "name": "自定义角色与权限",
        "idempotency": "role_name",
        "columns": [
            _col("role_name", "角色名", required=True, example="巡检主管"),
            _col("description", "描述", required=False, example="外勤管理"),
            _col("scopes", "权限 scope", required=True,
                 example="survey.read;geo.read",
                 doc="分号分隔；只能使用已注册 permission bundle"),
        ],
    },
    "memberships_v1": {
        "name": "用户-角色-客户授权",
        "idempotency": "username + role + customer_id + project_id",
        "columns": [
            _col("username", "用户名", required=True, example="alice"),
            _col("role", "角色名", required=True, example="analyst"),
            _col("customer_id", "客户ID", required=False,
                 example="cust-demo", doc="留空=角色级（无客户限定）"),
            _col("project_id", "项目ID", required=False, example=""),
        ],
    },
    "survey_definition_v1": {
        "name": "问卷定义（草稿）",
        "idempotency": "survey_name",
        "columns": [
            _col("survey_name", "问卷名称", required=True,
                 example="门店巡检问卷"),
            _col("description", "说明", required=False, example="月度巡检"),
            _col("pages", "页数", type_="int", required=False, example="1"),
        ],
    },
    "survey_questions_v1": {
        "name": "问卷题目",
        "idempotency": "survey_name + question_id",
        "columns": [
            _col("survey_name", "问卷名称", required=True,
                 example="门店巡检问卷"),
            _col("question_id", "题目ID", required=True, example="q1"),
            _col("qtype", "题型", required=True,
                 enum=["single_choice", "multi_choice", "text", "number",
                       "date", "rating", "matrix", "photo", "description"],
                 example="single_choice"),
            _col("title", "题目标题", required=True, example="货架是否整洁"),
            _col("options", "选项", required=False, example="是;否",
                 doc="竖线或分号分隔；photo 题此处填最少张数"),
            _col("required", "必填", type_="bool", required=False,
                 example="true"),
            _col("score", "分值", type_="float", required=False, example="5"),
            _col("dimension", "评分维度", required=False, example="陈列"),
        ],
    },
    "survey_logic_v1": {
        "name": "问卷跳题逻辑",
        "idempotency": "survey_name + from_question + to_question",
        "columns": [
            _col("survey_name", "问卷名称", required=True,
                 example="门店巡检问卷"),
            _col("from_question", "来源题目ID", required=True, example="q1"),
            _col("op", "运算符", required=True,
                 enum=["eq", "ne", "gt", "ge", "lt", "le"], example="eq"),
            _col("value", "比较值", required=True, example="否"),
            _col("to_question", "跳转目标题目ID", required=True,
                 example="q3"),
        ],
    },
    "route_constraints_v1": {
        "name": "路线规划约束",
        "idempotency": "customer_id + preset_name",
        "columns": [
            _col("customer_id", "客户ID", required=True, example="cust-demo"),
            _col("preset_name", "约束预设名", required=True,
                 example="华东日常巡检"),
            _col("max_km_per_day", "单日最大里程(km)", type_="float",
                 required=False, example="120"),
            _col("time_windows", "时间窗", required=False,
                 example="09:00-12:00;13:00-18:00"),
            _col("capacity", "单日任务容量", type_="int", required=False,
                 example="12"),
            _col("cost_per_km", "每公里成本", type_="float", required=False,
                 example="2.0"),
            _col("hard_isolate_projects", "多项目硬隔离", type_="bool",
                 required=False, example="true"),
        ],
    },
    "usage_rate_cards_v1": {
        "name": "Usage 价目卡",
        "idempotency": "rate_card_id（整卡生成新版本，历史账单不重算）",
        "columns": [
            _col("rate_card_id", "价目卡ID", required=True,
                 example="rc-standard"),
            _col("name", "价目卡名称", required=False, example="标准价目"),
            _col("unit", "计量单位", required=True,
                 example="recognition_photo",
                 doc="如 recognition_photo / model_compute_ms / "
                     "agent_output_token / survey_response / field_visit"),
            _col("price", "单价", type_="float", required=True,
                 example="0.5"),
            _col("currency", "币种", required=False, example="CNY"),
        ],
    },
    "knowledge_documents_v1": {
        "name": "知识库文档登记",
        "idempotency": "kb_name + title + version",
        "columns": [
            _col("kb_name", "知识库名称", required=True, example="巡检手册"),
            _col("customer_id", "客户范围", required=False, example="",
                 doc="留空=全平台"),
            _col("title", "文档标题", required=True, example="门店陈列规范"),
            _col("source", "文件/URL", required=False,
                 example="docs/kb/store-spec.md"),
            _col("version", "版本", required=False, example="v1"),
            _col("expires_at", "失效时间", type_="date", required=False,
                 example=""),
        ],
    },
}

# 状态机（03 文档 §导入状态机）
STATUSES = ("uploaded", "parsed", "mapped", "validated", "dry_run_passed",
            "awaiting_approval", "committed", "reconciled",
            "parse_failed", "validation_failed", "partial_failed",
            "compensated")


class ImportCenter:
    def __init__(self, store: Any, *, iam: Any, master: Any,
                 survey: Any, field_ops: Any, finance: Any) -> None:
        self.store = store
        self.iam = iam
        self.master = master
        self.survey = survey
        self.field_ops = field_ops
        self.finance = finance

    # ---------- 模板 ----------

    def list_templates(self) -> list[dict]:
        out = []
        for tid, t in TEMPLATES.items():
            out.append({"template_id": tid, "name": t["name"],
                        "idempotency": t["idempotency"],
                        "note": t.get("note", ""),
                        "columns": t["columns"]})
        return out

    def render_template(self, template_id: str,
                        fmt: str) -> tuple[bytes, str]:
        """返回 (bytes, filename)；模板含字段说明/样例/幂等键，
        且必须能被同一系统重新解析（round-trip 测试强制）。"""
        t = TEMPLATES.get(template_id)
        if t is None:
            raise ImportError_(f"模板不存在: {template_id}")
        fmt = fmt.lower().lstrip(".")
        if fmt not in ("csv", "xlsx"):
            raise ImportError_(f"模板格式不支持: {fmt}（仅 csv/xlsx）")
        header = [c["field"] for c in t["columns"]]
        sample = [c["example"] for c in t["columns"]]
        if fmt == "csv":
            buf = io.StringIO()
            buf.write(f"# 模板: {t['name']}（{template_id}）\n")
            buf.write(f"# 幂等键: {t['idempotency']}\n")
            for c in t["columns"]:
                req = "必填" if c["required"] else "可选"
                enum = f"；枚举={c['enum'][1:]}" if c["enum"] else ""
                doc = f"；{c['doc']}" if c["doc"] else ""
                buf.write(f"# {c['field']}（{c['label']}）：{c['type']}"
                          f"，{req}{enum}{doc}\n")
            w = csv.writer(buf)
            w.writerow(header)
            w.writerow(sample)
            data = buf.getvalue().encode("utf-8-sig")
            return data, f"{template_id}.csv"
        wb = Workbook()
        ws = wb.active
        ws.title = "data"
        ws.append(header)
        ws.append(sample)
        ws2 = wb.create_sheet("说明")
        ws2.append(["字段", "标签", "类型", "必填", "枚举", "样例", "说明"])
        for c in t["columns"]:
            ws2.append([c["field"], c["label"], c["type"],
                        "是" if c["required"] else "否",
                        "|".join(c["enum"][1:]) if c["enum"] else "",
                        c["example"], c["doc"] or t["idempotency"]])
        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue(), f"{template_id}.xlsx"

    # ---------- 授权（OSV5 指令 5.2/六；fail-closed） ----------

    def _platform_actor(self, actor: str, session_role: str) -> bool:
        if session_role == "admin":
            return True
        roles = self.iam.roles_of(actor)
        return "owner" in roles or "platform_admin" in roles

    def authorize_template(self, actor: str, session_role: str,
                           template_id: str) -> None:
        if template_id not in TEMPLATE_SCOPE:
            raise ImportError_(f"模板不存在: {template_id}")
        if self._platform_actor(actor, session_role):
            return
        scope = TEMPLATE_SCOPE[template_id]
        if not self.iam.authorize(actor, scope):
            raise ImportAuthError(
                f"IMPORT_PERMISSION_DENIED: 模板 {template_id} 需要"
                f" {scope}")

    def authorize_customers(self, actor: str, session_role: str,
                            customer_ids: list[str]) -> None:
        """逐客户授权：任一客户无权 → 整批拒绝（指令 5.2）。"""
        if not customer_ids or self._platform_actor(actor, session_role):
            return
        visible = self.iam.visible_customers(actor)
        if visible is None:
            return
        denied = sorted(set(customer_ids) - set(visible))
        if denied:
            raise ImportAuthError(
                "IMPORT_CUSTOMER_DENIED: 无权客户 "
                f"{denied[:5]}（整批 fail-closed）")

    def authorize_batch(self, actor: str, session_role: str,
                        batch: dict, *, write: bool = False) -> None:
        """批次作用域授权：平台角色放行；否则批次客户集 ⊆ 调用者
        可见客户；无客户作用域的全局批次需 data.import.audit。"""
        if actor == batch.get("actor"):
            return  # 创建者本人
        if self._platform_actor(actor, session_role):
            return
        cids = [r["customer_id"] for r in
                self.customer_scopes(batch["batch_id"])]
        if cids:
            visible = self.iam.visible_customers(actor)
            if visible is None:
                return
            if set(cids) - set(visible):
                raise ImportAuthError(
                    "IMPORT_BATCH_SCOPE_DENIED: 批次客户作用域超出"
                    "授权范围")
            if write and not self.iam.authorize(actor,
                    TEMPLATE_SCOPE.get(batch["template_id"],
                                       "master.manage")):
                raise ImportAuthError(
                    "IMPORT_PERMISSION_DENIED: 无写入权限")
            return
        if not self.iam.authorize(actor, "data.import.audit"):
            raise ImportAuthError(
                "IMPORT_BATCH_SCOPE_DENIED: 全局批次需"
                " data.import.audit")

    def customer_scopes(self, batch_id: str) -> list[dict]:
        rows = self.store._conn.execute(
            "SELECT customer_id, project_id, scope_source,"
            " authorization_decision FROM"
            " import_batch_customer_scope_v1 WHERE batch_id=?"
            " ORDER BY customer_id", (batch_id,)).fetchall()
        return [dict(r) for r in rows]

    # ---------- 上传 / 解析 ----------

    def upload(self, *, template_id: str, filename: str, data: bytes,
               actor: str, session_role: str = "admin",
               test_run_id: str = "", correlation_id: str = "") -> dict:
        """OSV5：批次 = 冻结执行上下文。授权先于落库（整批
        fail-closed）；携带 test_run_id 时同事务写 uat_fixture。"""
        if template_id not in TEMPLATES:
            raise ImportError_(f"模板不存在: {template_id}")
        self.authorize_template(actor, session_role, template_id)
        batch_id = _new_id("imp")
        fhash = hashlib.sha256(data).hexdigest()
        fmt = "xlsx" if filename.lower().endswith(
            (".xlsx", ".xlsm")) else "csv"
        # 作用域解析（fail-closed：archived/不存在 test_run 拒绝）
        data_scope = "operational"
        if test_run_id:
            from .scope import ScopeResolver
            ScopeResolver(self.store).assert_test_run_current(test_run_id)
            data_scope = "uat_fixture"
        try:
            header, rows = self._parse(fmt, data)
        except ImportError_:
            self._save_batch(batch_id, template_id, filename, fmt, fhash,
                             actor, "parse_failed", 0,
                             {"header": [], "rows": []}, [], {},
                             data_scope=data_scope,
                             test_run_id=test_run_id,
                             correlation_id=correlation_id)
            raise
        # 行内容随批次落库（dry-run/提交可重读；不依赖临时文件）
        mapping: dict[str, Any] = {"header": header, "rows": rows}
        mapped_cols = self._map_columns(TEMPLATES[template_id], header)
        mapping["column_map"] = mapped_cols
        missing = [c["field"] for c in TEMPLATES[template_id]["columns"]
                   if c["required"] and c["field"] not in mapped_cols]
        # 批次客户作用域推导 + 逐客户整批授权（落库前）
        customers = self._customers_of(template_id, mapped_cols, rows)
        self.authorize_customers(actor, session_role, customers)
        status = "parsed" if not missing else "parse_failed"
        self._save_batch(batch_id, template_id, filename, fmt, fhash,
                         actor, status, len(rows), mapping,
                         [{"row": 0, "error": f"缺少必填列: {missing}"}]
                         if missing else [], {},
                         data_scope=data_scope,
                         test_run_id=test_run_id,
                         correlation_id=correlation_id,
                         customers=customers)
        return self.batch_dto(self._must(batch_id))

    def _customers_of(self, template_id: str,
                      mapping: dict[str, int],
                      rows: list[list[str]]) -> list[str]:
        col = TEMPLATE_CUSTOMER_COL.get(template_id)
        if col is None:
            return []
        idx = mapping.get(col)
        if idx is None:
            return []
        out: list[str] = []
        for row in rows:
            v = row[idx].strip() if idx < len(row) else ""
            if v and v not in out:
                out.append(v)
        return out

    def _parse(self, fmt: str, data: bytes
               ) -> tuple[list[str], list[list[str]]]:
        if fmt == "csv":
            text = data.decode("utf-8-sig", errors="replace")
            lines = [l for l in text.splitlines()
                     if l.strip() and not l.lstrip().startswith("#")]
            if not lines:
                raise ImportError_("CSV 为空（或只有注释行）")
            reader = csv.reader(io.StringIO("\n".join(lines)))
            all_rows = [r for r in reader if any(c.strip() for c in r)]
            return all_rows[0], all_rows[1:]
        try:
            wb = load_workbook(io.BytesIO(data), read_only=True,
                               data_only=True)
        except Exception as e:
            raise ImportError_(f"XLSX 解析失败: {e}")
        ws = wb[wb.sheetnames[0]]
        all_rows = [[("" if c is None else str(c)).strip() for c in row]
                    for row in ws.iter_rows(values_only=True)]
        all_rows = [r for r in all_rows if any(c for c in r)]
        if not all_rows:
            raise ImportError_("XLSX 第一个工作表为空")
        return all_rows[0], all_rows[1:]

    # ---------- dry-run / 校验 ----------

    def dry_run(self, batch_id: str, *, actor: str = "",
                session_role: str = "admin") -> dict:
        b = self._must(batch_id)
        self._assert_batch_writable(b)
        if actor:
            self.authorize_batch(actor, session_role, b, write=True)
        if b["status"] == "parse_failed":
            raise ImportError_("批次解析失败，无法 dry-run（请重新上传）")
        header, rows = self._reread(b)
        t = TEMPLATES[b["template_id"]]
        mapping = self._map_columns(t, header)
        errors: list[dict] = []
        plan = {"insert": 0, "skip": 0, "update": 0, "conflict": 0}
        seen_keys: set[str] = set()
        for i, row in enumerate(rows, start=2):  # 行号从 2（1 是表头）
            rec, row_errs = self._validate_row(t, mapping, row, i)
            if row_errs:
                errors.extend(row_errs)
                continue
            action, reason = self._classify(b["template_id"], rec,
                                            seen_keys)
            plan[action] += 1
            if action == "conflict":
                errors.append({"row": i, "error": f"冲突：{reason}"})
        status = "dry_run_passed" if not errors else "validation_failed"
        self._update_batch(batch_id, status=status,
                           mapping={**b["mapping"], "column_map": mapping},
                           dry_run={"plan": plan, "rows": len(rows)},
                           errors=errors)
        return self.batch_dto(self._must(batch_id))

    def _assert_batch_writable(self, b: dict) -> None:
        """OSV51 C-1：作用域写冻结守卫（不可绕过的唯一强制点）。

        data_scope ∈ {quarantine, archived} 或 visibility=='history' 的
        批次一律禁止 dry-run/commit/重放：
        - quarantine：隔离裁决闭环（03-QUARANTINE-STATE-MACHINE.md）之前
          不得产生任何写入；quarantine dry-run 同样被拦（会覆写原批次
          dry_run_json 证据）。只读分析产物将来进裁决证据（契约 C-3
          adjudication evidence：追加式、独立表，后续波次实现），不改
          原批次行；
        - archived/history：已归档证据不得再次操作（OSV5 原语义并入，
          已终态历史批次重放同码）。
        稳定错误码 IMPORT_BATCH_WRITE_BLOCKED → API 映射 409，detail
        原样携码。守卫只读 DB 行（b 来自 _must），请求参数无法覆盖。
        """
        if b.get("data_scope") in ("quarantine", "archived") or \
                b.get("visibility") == "history":
            raise ImportError_(
                "IMPORT_BATCH_WRITE_BLOCKED: "
                f"data_scope={b.get('data_scope')} "
                f"visibility={b.get('visibility')}；"
                "隔离/归档批次不得执行 dry-run/commit，"
                "请走隔离区裁决流程")

    def _guard_active(self, b: dict) -> None:
        """OSV5 旧入口：统一委托 C-1 守卫（语义并入，保留兼容）。"""
        self._assert_batch_writable(b)

    def _map_columns(self, t: dict, header: list[str]) -> dict[str, int]:
        mapping: dict[str, int] = {}
        by_field = {c["field"]: i for i, c in enumerate(t["columns"])}
        for i, h in enumerate(header):
            h = (h or "").strip()
            if h in by_field:
                mapping[h] = i
                continue
            for c in t["columns"]:
                if c["label"] == h and c["field"] not in mapping:
                    mapping[c["field"]] = i
        return mapping

    def _validate_row(self, t: dict, mapping: dict[str, int],
                      row: list[str], row_no: int
                      ) -> tuple[dict, list[dict]]:
        rec: dict[str, Any] = {}
        errs: list[dict] = []
        for c in t["columns"]:
            f = c["field"]
            idx = mapping.get(f)
            raw = row[idx].strip() if idx is not None and idx < len(
                row) else ""
            if not raw:
                if c["required"]:
                    errs.append({"row": row_no,
                                 "error": f"必填字段缺失: {f}（{c['label']}）"})
                rec[f] = None
                continue
            if c["enum"] and raw not in c["enum"]:
                errs.append({"row": row_no,
                             "error": f"{f} 值不在枚举内: {raw}"
                                      f"（允许 {c['enum'][1:]}）"})
            if c["type"] in ("int", "float"):
                try:
                    rec[f] = int(raw) if c["type"] == "int" else float(raw)
                except ValueError:
                    errs.append({"row": row_no,
                                 "error": f"{f} 不是数字: {raw}"})
                    rec[f] = None
            elif c["type"] == "bool":
                if raw.lower() in ("true", "1", "是", "yes"):
                    rec[f] = True
                elif raw.lower() in ("false", "0", "否", "no"):
                    rec[f] = False
                else:
                    errs.append({"row": row_no,
                                 "error": f"{f} 不是布尔值: {raw}"})
                    rec[f] = None
            elif c["type"] == "date":
                try:
                    datetime.strptime(raw[:10], "%Y-%m-%d")
                    rec[f] = raw[:10]
                except ValueError:
                    errs.append({"row": row_no,
                                 "error": f"{f} 日期格式应为 YYYY-MM-DD: {raw}"})
                    rec[f] = None
            else:
                rec[f] = raw
        return rec, errs

    def _classify(self, template_id: str, rec: dict,
                  seen_keys: set[str]) -> tuple[str, str]:
        """按自然键分类 insert/skip/conflict（幂等提交依据）。"""
        def dup(key: str) -> bool:
            if key in seen_keys:
                return True
            seen_keys.add(key)
            return False

        conn = self.store._conn
        if template_id == "customers_v1":
            k = rec["customer_id"]
            if dup(f"c:{k}"):
                return "conflict", "文件内 customer_id 重复"
            if self.master.get_customer(k):
                return "skip", ""
            return "insert", ""
        if template_id == "projects_v1":
            k = rec["project_id"]
            if dup(f"p:{k}"):
                return "conflict", "文件内 project_id 重复"
            if self.master.get_project(k):
                return "skip", ""
            if self.master.get_customer(rec["customer_id"]) is None:
                return "conflict", f"客户不存在: {rec['customer_id']}"
            return "insert", ""
        if template_id == "skus_v1":
            k = rec["sku_id"]
            if dup(f"s:{k}"):
                return "conflict", "文件内 sku_id 重复"
            if self.master.get_sku(k):
                return "skip", ""
            return "insert", ""
        if template_id == "stores_addresses_v1":
            k = f"{rec['customer_id']}|{rec['store_name']}"
            if dup(f"a:{k}"):
                return "conflict", "文件内门店重复"
            hit = conn.execute(
                "SELECT 1 FROM geo_address_v1 WHERE customer_id=?"
                " AND raw LIKE ? LIMIT 1",
                (rec["customer_id"],
                 f"%{rec['store_name']}%")).fetchone()
            return ("skip", "") if hit else ("insert", "")
        if template_id == "employees_v1":
            k = f"{rec['customer_id']}|{rec['name']}"
            if dup(f"e:{k}"):
                return "conflict", "文件内员工重复"
            hit = conn.execute(
                "SELECT 1 FROM geo_employee_v1 WHERE customer_id=?"
                " AND name=? LIMIT 1",
                (rec["customer_id"], rec["name"])).fetchone()
            return ("skip", "") if hit else ("insert", "")
        if template_id == "users_v1":
            k = rec["username"]
            if dup(f"u:{k}"):
                return "conflict", "文件内用户名重复"
            if self.iam.get_principal_by_username(k):
                return "skip", ""
            return "insert", ""
        if template_id == "roles_permissions_v1":
            k = rec["role_name"]
            if dup(f"r:{k}"):
                return "conflict", "文件内角色重复"
            if conn.execute("SELECT 1 FROM iam_role_v1 WHERE name=?",
                            (k,)).fetchone():
                return "skip", ""
            return "insert", ""
        if template_id == "memberships_v1":
            k = "|".join(str(rec.get(x) or "") for x in
                         ("username", "role", "customer_id", "project_id"))
            if dup(f"m:{k}"):
                return "conflict", "文件内授权重复"
            if self.iam.get_principal_by_username(
                    rec["username"]) is None:
                return "conflict", f"用户不存在: {rec['username']}"
            return "insert", ""
        if template_id in ("survey_definition_v1", "survey_questions_v1",
                           "survey_logic_v1"):
            k = f"{template_id}:{rec['survey_name']}" + (
                f":{rec.get('question_id')}" if rec.get("question_id")
                else "")
            if dup(k):
                return "conflict", "文件内问卷行重复"
            return "insert", ""
        if template_id == "route_constraints_v1":
            k = f"{rec['customer_id']}|{rec['preset_name']}"
            if dup(f"rc:{k}"):
                return "conflict", "文件内预设重复"
            return "insert", ""
        if template_id == "usage_rate_cards_v1":
            k = f"{rec['rate_card_id']}|{rec['unit']}"
            if dup(f"rc2:{k}"):
                return "conflict", "文件内价目行重复"
            return "insert", ""
        if template_id == "knowledge_documents_v1":
            k = f"{rec['kb_name']}|{rec['title']}|{rec.get('version') or 'v1'}"
            if dup(f"kb:{k}"):
                return "conflict", "文件内文档重复"
            return "insert", ""
        return "insert", ""

    # ---------- 提交 ----------

    def commit(self, batch_id: str, *, actor: str,
               session_role: str = "admin") -> dict:
        b = self._must(batch_id)
        self._assert_batch_writable(b)
        self.authorize_batch(actor, session_role, b, write=True)
        if b["status"] not in ("dry_run_passed", "committed"):
            raise ImportError_(
                f"只有 dry_run_passed 可提交（当前 {b['status']}）；"
                "请先执行 dry-run")
        header, rows = self._reread(b)
        t = TEMPLATES[b["template_id"]]
        mapping = self._map_columns(t, header)
        stats = {"inserted": 0, "skipped": 0, "failed": 0}
        errors: list[dict] = []
        receipts: list[dict] = []
        seen_keys: set[str] = set()
        pending_survey_rows: list[dict] = []
        for i, row in enumerate(rows, start=2):
            rec, row_errs = self._validate_row(t, mapping, row, i)
            if row_errs:
                stats["failed"] += 1
                errors.extend(row_errs)
                continue
            action, _reason = self._classify(b["template_id"], rec,
                                             seen_keys)
            if action in ("skip", "conflict") and action == "skip":
                stats["skipped"] += 1
                continue
            try:
                out = self._commit_row(b["template_id"], rec, actor,
                                       pending_survey_rows, receipts)
                if out is not None:
                    receipts.append(out)
                stats["inserted"] += 1
            except Exception as e:
                stats["failed"] += 1
                errors.append({"row": i, "error": str(e)[:300]})
        # 问卷/价目类：按名称聚合后批量写入
        if b["template_id"] in ("survey_questions_v1", "survey_logic_v1",
                                "usage_rate_cards_v1"):
            try:
                self._flush_survey_rows(b["template_id"],
                                        pending_survey_rows, actor)
            except Exception as e:
                stats["failed"] += 1
                errors.append({"row": 0, "error": f"批量写入失败: {e}"})
        status = "committed" if stats["failed"] == 0 else "partial_failed"
        # OSV51 C-2：落库回执先脱敏；明文只存在于本次响应（内存）。
        once_receipts = [dict(r) for r in receipts[:50]]
        commit_result = {"stats": stats,
                         "receipts": redact_secrets(receipts[:50])}
        # OSV5：fixture 批次的导入对象继承批次作用域（同事务）。
        if b.get("data_scope") == "uat_fixture" and b.get("test_run_id"):
            self._inherit_batch_scope(b, receipts)
        self._update_batch(batch_id, status=status, errors=errors,
                           commit=commit_result)
        # OSV51 C-3：release revision 提交成功 → 源 quarantine 批次
        # superseded_by_new_batch（条件 UPDATE，仅 release_approved 时）。
        if b.get("source") == "quarantine_release" and \
                b.get("correlation_id"):
            self.store._conn.execute(
                "UPDATE quarantine_adjudication_v1 SET"
                " state='superseded_by_new_batch', version=version+1,"
                " updated_at=? WHERE batch_id=? AND"
                " state='release_approved'",
                (_now(), b["correlation_id"]))
            self.store._conn.commit()
        # 证据 + 审计（原文件 hash/actor/结果留痕）
        try:
            self.store.insert_evidence_bundle(
                evidence_id=_new_id("evid"), kind="import_batch",
                source_uri=f"import_batch:{batch_id}",
                content_type="text/csv", producer="import_center",
                input_hash=b["file_hash"],
                config_version=b["template_id"])
            self.iam.audit(actor, "import.committed",
                           f"import:{batch_id}",
                           {"template": b["template_id"],
                            "stats": stats},
                           customer_id="")
        except Exception:
            pass
        dto = self.batch_dto(self._must(batch_id))
        # OSV51 C-2：初始口令仅此一次随 commit 响应返回（不落库、
        # 不再被任何 GET/列表/重读路径返回）。
        dto["commit"] = {**(dto.get("commit") or {}),
                         "receipts": once_receipts}
        return dto

    def _inherit_batch_scope(self, b: dict, receipts: list) -> None:
        """fixture 批次提交对象作用域继承（自然键/回执键）。"""
        conn = self.store._conn
        trid, ds = b["test_run_id"], "uat_fixture"
        tpl = b["template_id"]
        if tpl in _SCOPE_INHERIT_NATURAL:
            table, idcol = _SCOPE_INHERIT_NATURAL[tpl]
            header, rows = self._reread(b)
            mapping = self._map_columns(TEMPLATES[tpl], header)
            idx = mapping.get(idcol)
            if idx is not None:
                keys = sorted({row[idx].strip() for row in rows
                               if idx < len(row) and row[idx].strip()})
                for k in keys:
                    conn.execute(
                        f"UPDATE {table} SET data_scope=?, test_run_id=?"
                        f" WHERE {idcol}=? AND COALESCE(test_run_id,'')"
                        "=''", (ds, trid, k))
        if tpl in _SCOPE_INHERIT_RECEIPT:
            table, idcol = _SCOPE_INHERIT_RECEIPT[tpl]
            ids = sorted({r[idcol] for r in receipts if r.get(idcol)})
            for oid in ids:
                conn.execute(
                    f"UPDATE {table} SET data_scope=?, test_run_id=?"
                    f" WHERE {idcol}=? AND COALESCE(test_run_id,'')=''",
                    (ds, trid, oid))
        conn.commit()

    def _commit_row(self, template_id: str, rec: dict, actor: str,
                    pending_survey: list, receipts: list
                    ) -> dict | None:
        if template_id == "customers_v1":
            self.master.create_customer(
                customer_id=rec["customer_id"], name=rec["name"],
                retention_policy=rec.get("retention_policy") or "",
                created_by=actor)
            return None
        if template_id == "projects_v1":
            self.master.create_project(
                project_id=rec["project_id"],
                customer_id=rec["customer_id"], name=rec["name"],
                budget={"total": rec.get("budget") or 0,
                        "owner": rec.get("owner") or "",
                        "start": rec.get("start_date") or "",
                        "end": rec.get("end_date") or ""},
                created_by=actor)
            return None
        if template_id == "skus_v1":
            self.master.create_sku(
                sku_id=rec["sku_id"],
                canonical_name=rec["canonical_name"],
                brand=rec.get("brand") or "",
                category=rec.get("category") or "",
                volume=rec.get("volume") or "",
                barcode=rec.get("barcode") or "",
                package_version=rec.get("package_version") or "v1",
                valid_from=rec.get("valid_from"),
                valid_to=rec.get("valid_to"), created_by=actor)
            for alias in (rec.get("aliases") or "").replace("|", ";")\
                    .split(";"):
                if alias.strip():
                    self.master.add_alias(sku_id=rec["sku_id"],
                                          alias=alias.strip(),
                                          actor=actor)
            return None
        if template_id == "stores_addresses_v1":
            raw = f"{rec['store_name']}｜{rec['raw_address']}"
            addr = self.field_ops.add_address(
                customer_id=rec["customer_id"], raw=raw, actor=actor)
            if rec.get("lat") is not None and rec.get("lng") is not None:
                # 用户自带坐标：登记为人工确认（source=import），
                # 不伪造地理编码候选
                self.store._conn.execute(
                    "UPDATE geo_address_v1 SET status='confirmed',"
                    " chosen_json=?, confidence=1.0, verified_by=?"
                    " WHERE address_id=?",
                    (json.dumps({"lat": rec["lat"], "lng": rec["lng"],
                                 "source": "import",
                                 "coord_system": rec.get("coord_system")
                                 or "wgs84",
                                 "time_window": rec.get("time_window")
                                 or "", "region": rec.get("region") or ""},
                                ensure_ascii=False),
                     actor, addr["address_id"]))
                self.store._conn.commit()
            return {"address_id": addr["address_id"]}
        if template_id == "employees_v1":
            emp = self.field_ops.add_employee(
                customer_id=rec["customer_id"], name=rec["name"],
                skills=[s for s in (rec.get("skills") or "")
                        .replace("|", ";").split(";") if s.strip()],
                vehicle=rec.get("vehicle") or "")
            return {"employee_id": emp["employee_id"]}
        if template_id == "users_v1":
            import secrets
            temp_pw = "Init-" + secrets.token_hex(16)
            self.iam.create_principal(
                kind=rec.get("kind") or "user", username=rec["username"],
                display_name=rec.get("display_name") or "",
                password=temp_pw, created_by=actor)
            # OSV51 C-2：一次性初始口令只在 commit 当次响应中返回；
            # 落库 receipts 一律经 redact_secrets 脱敏（DB 只存哈希）。
            return {"username": rec["username"],
                    "initial_password_once": temp_pw}
        if template_id == "roles_permissions_v1":
            scopes = [s.strip() for s in (rec["scopes"] or "")
                      .replace("|", ";").split(";") if s.strip()]
            return self.iam.create_custom_role(
                name=rec["role_name"],
                description=rec.get("description") or "",
                scopes=scopes, created_by=actor)
        if template_id == "memberships_v1":
            return self.iam.grant(username=rec["username"],
                                  role=rec["role"],
                                  customer_id=rec.get("customer_id") or "",
                                  project_id=rec.get("project_id") or "",
                                  granted_by=actor)
        if template_id == "survey_definition_v1":
            hit = next((s for s in self.survey.list_surveys()
                        if s["name"] == rec["survey_name"]), None)
            if hit is not None:
                return None  # 幂等：同名问卷已存在
            self.survey.create_draft(name=rec["survey_name"],
                                     spec={"pages": int(
                                         rec.get("pages") or 1),
                                         "description":
                                         rec.get("description") or "",
                                         "questions": []},
                                     actor=actor)
            return None
        if template_id in ("survey_questions_v1", "survey_logic_v1"):
            pending_survey.append(rec)
            return None
        if template_id == "route_constraints_v1":
            constraints = {k: rec.get(k) for k in
                           ("max_km_per_day", "time_windows", "capacity",
                            "cost_per_km", "hard_isolate_projects")}
            self.store._conn.execute(
                "INSERT INTO route_constraint_preset_v1 (preset_id,"
                " customer_id, name, constraints_json, created_by,"
                " created_at) VALUES (?,?,?,?,?,?)",
                (_new_id("rcp"), rec["customer_id"], rec["preset_name"],
                 json.dumps(constraints, ensure_ascii=False), actor,
                 _now()))
            self.store._conn.commit()
            return None
        if template_id == "usage_rate_cards_v1":
            # 聚合在 commit 外层按 rate_card_id 成卡：这里逐行写临时
            pending_survey.append(rec)  # 复用暂存列表
            return None
        if template_id == "knowledge_documents_v1":
            self.store._conn.execute(
                "INSERT INTO knowledge_document_v1 (doc_id, kb_name,"
                " customer_id, title, source, version, expires_at,"
                " status, created_by, created_at)"
                " VALUES (?,?,?,?,?,?,?,?,?,?)",
                (_new_id("kbdoc"), rec["kb_name"],
                 rec.get("customer_id") or "", rec["title"],
                 rec.get("source") or "", rec.get("version") or "v1",
                 rec.get("expires_at"), "draft", actor, _now()))
            self.store._conn.commit()
            return None
        raise ImportError_(f"未实现提交处理: {template_id}")

    def _flush_survey_rows(self, template_id: str, rows: list[dict],
                           actor: str) -> None:
        if template_id == "usage_rate_cards_v1":
            # 价目卡按 rate_card_id 聚合：整卡生成新版本（幂等：
            # 同批次重复提交时 skip 已在 classify 处理）
            cards: dict[str, dict] = {}
            for rec in rows:
                c = cards.setdefault(rec["rate_card_id"], {
                    "name": rec.get("name") or rec["rate_card_id"],
                    "lines": [], "currency": rec.get("currency") or "CNY"})
                c["lines"].append({"unit": rec["unit"],
                                   "price": rec["price"]})
            for rc_id, c in cards.items():
                try:
                    self.finance.get_rate_card(rc_id)
                    exists = True
                except Exception:
                    exists = False
                if not exists:
                    self.store._conn.execute(
                        "INSERT INTO fin_rate_card_v1 (rate_card_id,"
                        " version, name, lines_json, currency,"
                        " created_at) VALUES (?,?,?, ?,?,?)",
                        (rc_id, 1, c["name"],
                         json.dumps(c["lines"]), c["currency"], _now()))
                    self.store._conn.commit()
                else:
                    self.finance.new_rate_card_version(
                        rc_id, lines=c["lines"], actor=actor)
            return
        by_survey: dict[str, list[dict]] = {}
        for rec in rows:
            by_survey.setdefault(rec["survey_name"], []).append(rec)
        for name, recs in by_survey.items():
            hit = next((s for s in self.survey.list_surveys()
                        if s["name"] == name), None)
            if hit is None:
                # 问卷不存在时自动建 draft（导入宽容）
                hit = self.survey.create_draft(
                    name=name, spec={"questions": []}, actor=actor)
            spec = json.loads(json.dumps(hit.get("spec") or {}))
            questions = spec.setdefault("questions", [])
            if template_id == "survey_questions_v1":
                for rec in recs:
                    q = {"id": rec["question_id"],
                         "type": rec["qtype"], "title": rec["title"],
                         "required": bool(rec.get("required"))}
                    if rec.get("options"):
                        q["options"] = [o for o in str(rec["options"])
                                        .replace("|", ";").split(";")
                                        if o.strip()]
                    if rec.get("score") is not None:
                        q["score"] = rec["score"]
                    if rec.get("dimension"):
                        q["dimension"] = rec["dimension"]
                    questions = [x for x in questions
                                 if x.get("id") != q["id"]] + [q]
            else:  # survey_logic_v1
                logic = spec.setdefault("skip_logic", [])
                for rec in recs:
                    rule = {"from": rec["from_question"],
                            "op": rec["op"], "value": rec["value"],
                            "to": rec["to_question"]}
                    logic = [x for x in logic if x != rule] + [rule]
                spec["skip_logic"] = logic
            spec["questions"] = questions
            self.survey.update_draft(hit["survey_id"], spec=spec)

    # ---------- 查询（OSV5：DTO 白名单 + 作用域过滤） ----------

    _DTO_KEYS = ("batch_id", "template_id", "filename", "file_format",
                 "file_hash", "status", "actor", "row_count",
                 "data_scope", "test_run_id", "visibility",
                 "archived_at", "source", "correlation_id",
                 "created_at", "updated_at")

    def batch_dto(self, b: dict) -> dict:
        """显式响应 DTO：绝不返回 mapping_json/dry_run_json/
        error_report_json/commit_json 等原始 payload（指令 P0-004）。"""
        d = {k: b.get(k) for k in self._DTO_KEYS}
        d["customer_scopes"] = [
            {"customer_id": r["customer_id"],
             "project_id": r["project_id"],
             "authorization_decision": r["authorization_decision"]}
            for r in self.customer_scopes(b["batch_id"])]
        d["dry_run"] = {k: b.get("dry_run", {}).get(k)
                        for k in ("plan", "rows")}
        d["errors"] = [dict(e) for e in (b.get("errors") or [])]
        d["error_count"] = len(d["errors"])
        c = b.get("commit") or {}
        d["commit"] = {"stats": c.get("stats"),
                       "receipts": c.get("receipts")}
        # OSV51 C-3：隔离批次附带裁决状态（UI 裁决面板的数据源）。
        if b.get("data_scope") == "quarantine":
            d["adjudication"] = self.adjudication_dto(b["batch_id"])
        # OSV51 C-4：全局模板标记（空客户范围时 UI 仅对全局模板显示
        # “全局”；有客户列模板空关联 → 未绑定/待裁决）。
        d["is_global_template"] = TEMPLATE_CUSTOMER_COL.get(
            b.get("template_id") or "") is None
        # OSV51 C-2：出口兜底——DTO 全量递归 secret 扫描（防存量行
        # 与任何新增字段的明文泄漏；commit 当次响应的明文回执由
        # commit() 在返回前覆盖，不经此路径）。
        return redact_secrets(d)

    def preview_rows(self, batch_id: str, *, limit: int = 50) -> dict:
        """原始行预览（仅创建者/data.import.audit，API 层鉴权）；
        显式脱敏标记 + 行数上限。"""
        b = self._must(batch_id)
        header, rows = self._reread(b)
        return {"batch_id": batch_id, "redacted": True,
                "header": header, "rows": rows[:limit],
                "truncated": len(rows) > limit,
                "total_rows": len(rows)}

    def list_batches(self, *, actor: str = "", session_role: str = "admin",
                     view: str = "operational",
                     include_fixture: bool = False) -> list[dict]:
        """OSV5 列表口径：默认 effective operational ∩ 调用者客户
        作用域；history/quarantine 仅平台/auditor；include_fixture
        需授权（fail-closed）。"""
        conn = self.store._conn
        platform = (not actor) or self._platform_actor(actor, session_role)
        if view == "quarantine":
            if not (platform or self.iam.authorize(
                    actor, "data.import.audit")):
                raise ImportAuthError(
                    "IMPORT_VIEW_DENIED: 隔离区仅管理员/审计可见")
            rows = conn.execute(
                "SELECT batch_id FROM import_batch_v1 WHERE"
                " COALESCE(data_scope,'operational')='quarantine'"
                " ORDER BY created_at DESC LIMIT 200").fetchall()
            return [self.batch_dto(self._must(r["batch_id"]))
                    for r in rows]
        if view == "history":
            if not (platform or (include_fixture and self.iam.authorize(
                    actor, "data.import.audit"))):
                raise ImportAuthError(
                    "IMPORT_VIEW_DENIED: 历史视图需授权")
            rows = conn.execute(
                "SELECT batch_id FROM import_batch_v1 WHERE"
                " COALESCE(visibility,'current')='history' OR"
                " COALESCE(data_scope,'operational') IN"
                " ('uat_fixture','demo_fixture')"
                " ORDER BY created_at DESC LIMIT 200").fetchall()
            return [self.batch_dto(self._must(r["batch_id"]))
                    for r in rows]
        if view == "mine":
            rows = conn.execute(
                "SELECT batch_id FROM import_batch_v1 WHERE actor=?"
                " ORDER BY created_at DESC LIMIT 200", (actor,)
            ).fetchall()
            return [self.batch_dto(self._must(r["batch_id"]))
                    for r in rows]
        # 默认：effective operational（排除 fixture/quarantine/history）
        rows = conn.execute(
            "SELECT batch_id FROM import_batch_v1 WHERE"
            " COALESCE(data_scope,'operational')='operational' AND"
            " COALESCE(visibility,'current')='current' AND"
            " COALESCE(test_run_id,'')=''"
            " ORDER BY created_at DESC LIMIT 200").fetchall()
        out = []
        visible = None if platform else self.iam.visible_customers(actor)
        for r in rows:
            dto = self.batch_dto(self._must(r["batch_id"]))
            cids = [c["customer_id"] for c in dto["customer_scopes"]]
            if visible is not None and cids and \
                    set(cids) - set(visible):
                continue  # 客户作用域外批次对非平台角色不可见
            out.append(dto)
        return out

    def get_batch(self, batch_id: str) -> dict:
        b = self._must(batch_id)
        return b

    def _must(self, batch_id: str) -> dict:
        row = self.store._conn.execute(
            "SELECT * FROM import_batch_v1 WHERE batch_id=?",
            (batch_id,)).fetchone()
        if row is None:
            raise ImportError_(f"导入批次不存在: {batch_id}")
        d = dict(row)
        d["mapping"] = json.loads(d["mapping_json"])
        d["dry_run"] = json.loads(d["dry_run_json"])
        d["errors"] = json.loads(d["error_report_json"])
        d["commit"] = json.loads(d["commit_json"])
        return d

    def _reread(self, b: dict) -> tuple[list[str], list[list[str]]]:
        """批次文件内容在上传时随 dry_run_json 保存（data_url 简化为
        内存重解析不可行，故 upload 时把行存入 mapping_json.rows）。"""
        rows = (b["mapping"] or {}).get("rows") or []
        header = (b["mapping"] or {}).get("header") or []
        return header, rows

    # ---------- 内部持久化 ----------

    def _save_batch(self, batch_id: str, template_id: str, filename: str,
                    fmt: str, fhash: str, actor: str, status: str,
                    row_count: int, mapping: dict, errors: list,
                    dry_run: dict, *, data_scope: str = "operational",
                    test_run_id: str = "", correlation_id: str = "",
                    customers: list[str] | None = None) -> None:
        conn = self.store._conn
        conn.execute(
            "INSERT INTO import_batch_v1 (batch_id, template_id, filename,"
            " file_format, file_hash, status, actor, row_count,"
            " mapping_json, dry_run_json, error_report_json, commit_json,"
            " created_at, updated_at, data_scope, test_run_id, source,"
            " correlation_id)"
            " VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)",
            (batch_id, template_id, filename, fmt, fhash, status, actor,
             row_count, json.dumps(mapping, ensure_ascii=False),
             json.dumps(dry_run, ensure_ascii=False),
             json.dumps(errors, ensure_ascii=False), "{}",
             _now(), _now(), data_scope, test_run_id, "import_center",
             correlation_id))
        # OSV5：多客户作用域关联（不得压成单 customer_id）。
        for cid in (customers or []):
            conn.execute(
                "INSERT OR IGNORE INTO import_batch_customer_scope_v1"
                " (batch_id, customer_id, project_id, scope_source,"
                " authorization_decision, created_at)"
                " VALUES (?,?,?,?,?,?)",
                (batch_id, cid, "", "row", "granted", _now()))
        conn.commit()

    def _update_batch(self, batch_id: str, *, status: str | None = None,
                      mapping: dict | None = None,
                      dry_run: dict | None = None,
                      errors: list | None = None,
                      commit: dict | None = None) -> None:
        b = self._must(batch_id)
        m = b["mapping"] if mapping is None else mapping
        self.store._conn.execute(
            "UPDATE import_batch_v1 SET status=?, mapping_json=?,"
            " dry_run_json=?, error_report_json=?, commit_json=?,"
            " updated_at=? WHERE batch_id=?",
            (status or b["status"], json.dumps(m, ensure_ascii=False),
             json.dumps(dry_run if dry_run is not None else b["dry_run"],
                        ensure_ascii=False),
             json.dumps(errors if errors is not None else b["errors"],
                        ensure_ascii=False),
             json.dumps(commit if commit is not None else b["commit"],
                        ensure_ascii=False),
             _now(), batch_id))
        self.store._conn.commit()

    # ---------- OSV51 C-3：隔离区人工裁决状态机 ----------
    # 状态集与契约：03-QUARANTINE-STATE-MACHINE.md。
    # 原则：原始导入证据不可修改；所有迁移 CAS（version 条件 UPDATE）；
    # release_to_operational 双人审批 + 新批次 revision（不原地改）。

    ADJ_STATES = ("quarantined", "retained_for_evidence",
                  "bound_to_test_run", "soft_discarded",
                  "release_requested", "release_approved",
                  "superseded_by_new_batch")

    def _adjudication_row(self, batch_id: str) -> dict:
        conn = self.store._conn
        row = conn.execute(
            "SELECT * FROM quarantine_adjudication_v1 WHERE batch_id=?",
            (batch_id,)).fetchone()
        if row:
            return dict(row)
        now = _now()
        conn.execute(
            "INSERT OR IGNORE INTO quarantine_adjudication_v1"
            " (batch_id, state, version, created_at, updated_at)"
            " VALUES (?,?,?,?,?)",
            (batch_id, "quarantined", 0, now, now))
        conn.commit()
        row = conn.execute(
            "SELECT * FROM quarantine_adjudication_v1 WHERE batch_id=?",
            (batch_id,)).fetchone()
        return dict(row)

    def adjudication_dto(self, batch_id: str) -> dict:
        return self._adjudication_row(batch_id)

    def adjudication_history(self, batch_id: str) -> list[dict]:
        rows = self.store._conn.execute(
            "SELECT kind, actor, detail_json, created_at FROM"
            " quarantine_adjudication_evidence_v1 WHERE batch_id=?"
            " ORDER BY id", (batch_id,)).fetchall()
        out = []
        for r in rows:
            d = {"kind": r["kind"], "actor": r["actor"],
                 "created_at": r["created_at"]}
            try:
                d["detail"] = json.loads(r["detail_json"] or "{}")
            except Exception:
                d["detail"] = {}
            out.append(d)
        return out

    def _create_release_revision(self, orig: dict, *, data_scope: str,
                                 test_run_id: str, actor: str) -> str:
        """创建新批次 revision（复制 mapping 行；原行不动）。"""
        conn = self.store._conn
        new_id = _new_id("imp")
        now = _now()
        conn.execute(
            "INSERT INTO import_batch_v1 (batch_id, template_id,"
            " filename, file_format, file_hash, status, actor,"
            " row_count, mapping_json, dry_run_json,"
            " error_report_json, commit_json, created_at, updated_at,"
            " data_scope, test_run_id, visibility, source,"
            " correlation_id) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)",
            (new_id, orig["template_id"],
             (orig.get("filename") or "") + ".release",
             orig.get("file_format") or "csv",
             orig.get("file_hash") or "", "uploaded", actor,
             orig.get("row_count") or 0,
             json.dumps(orig.get("mapping") or {}, ensure_ascii=False),
             "{}", "[]", "{}", now, now, data_scope, test_run_id,
             "current", "quarantine_release", orig["batch_id"]))
        if data_scope == "uat_fixture" and test_run_id:
            # 客户关联物化（单一关联源；来自 Test Run 注册表，
            # 不做名称猜测）
            tr = conn.execute(
                "SELECT customer_ids_json FROM uat_test_run_v1"
                " WHERE test_run_id=?", (test_run_id,)).fetchone()
            try:
                cids = json.loads((tr and tr["customer_ids_json"])
                                  or "[]")
            except Exception:
                cids = []
            for cid in cids:
                conn.execute(
                    "INSERT OR IGNORE INTO"
                    " import_batch_customer_scope_v1 (batch_id,"
                    " customer_id, project_id, scope_source,"
                    " authorization_decision, created_at)"
                    " VALUES (?,?,?,?,?,?)",
                    (new_id, cid, "", "adjudication_bind", "granted",
                     now))
        conn.commit()
        return new_id

    def adjudicate(self, batch_id: str, *, action: str, actor: str,
                   session_role: str = "admin", reason: str = "",
                   target_test_run_id: str = "",
                   version: int | None = None) -> dict:
        b = self._must(batch_id)
        if b.get("data_scope") != "quarantine":
            raise AdjudicationError(
                f"ADJUDICATION_NOT_QUARANTINE: 批次 {batch_id}"
                f" data_scope={b.get('data_scope')}，非隔离批次")
        if not (self._platform_actor(actor, session_role)
                or self.iam.authorize(actor, "data.import.audit")):
            raise ImportAuthError(
                "ADJUDICATION_PERMISSION_DENIED: 裁决需 platform 角色"
                " 或 data.import.audit")
        conn = self.store._conn
        cur = self._adjudication_row(batch_id)
        state = cur["state"]
        expected = cur["version"] if version is None else int(version)

        def _conflict():
            raise AdjudicationError(
                "ADJUDICATION_VERSION_CONFLICT: 裁决版本冲突"
                f"（请求 version={expected}，当前"
                f" version={cur['version']}）；请刷新后重试")

        def _invalid():
            raise AdjudicationError(
                f"ADJUDICATION_INVALID_TRANSITION: {state} 不允许"
                f" {action}")

        def _cas(new_state: str, **fields) -> dict:
            sets = ["state=?", "version=version+1", "updated_at=?"]
            vals: list = [new_state, _now()]
            for k in ("target_test_run_id", "revision_batch_id",
                      "requested_by", "requested_at", "approved_by",
                      "approved_at", "reason"):
                if k in fields:
                    sets.append(f"{k}=?")
                    vals.append(fields[k])
            vals.extend([batch_id, expected])
            rc = conn.execute(
                "UPDATE quarantine_adjudication_v1 SET"
                f" {', '.join(sets)} WHERE batch_id=? AND version=?",
                vals).rowcount
            conn.commit()
            return rc

        def _record(kind: str, detail: dict) -> None:
            conn.execute(
                "INSERT INTO quarantine_adjudication_evidence_v1"
                " (batch_id, kind, actor, detail_json, created_at)"
                " VALUES (?,?,?,?,?)",
                (batch_id, kind, actor,
                 json.dumps(detail, ensure_ascii=False), _now()))
            conn.commit()
            try:
                self.iam.audit(actor, f"import.quarantine.{kind}",
                               f"import:{batch_id}", detail,
                               customer_id="")
            except Exception:
                pass

        now = _now()
        if action == "retain":
            if state == "retained_for_evidence":
                return self.adjudication_dto(batch_id)  # 幂等
            if state not in ("quarantined", "release_requested"):
                _invalid()
            rc = _cas("retained_for_evidence", reason=reason)
            if rc == 0:
                _conflict()
            _record("retain", {"reason": reason})
        elif action == "soft_discard":
            if state == "soft_discarded":
                return self.adjudication_dto(batch_id)
            if state not in ("quarantined", "retained_for_evidence"):
                _invalid()
            rc = _cas("soft_discarded", reason=reason)
            if rc == 0:
                _conflict()
            _record("soft_discard", {"reason": reason})
        elif action == "bind_test_run":
            if state == "bound_to_test_run" and \
                    cur["target_test_run_id"] == target_test_run_id:
                return self.adjudication_dto(batch_id)
            if state not in ("quarantined", "retained_for_evidence"):
                _invalid()
            tr = conn.execute(
                "SELECT test_run_id FROM uat_test_run_v1"
                " WHERE test_run_id=?",
                (target_test_run_id,)).fetchone()
            if not tr:
                raise AdjudicationError(
                    "ADJUDICATION_TEST_RUN_NOT_FOUND: "
                    f"{target_test_run_id} 不在 Test Run 注册表")
            rev = self._create_release_revision(
                b, data_scope="uat_fixture",
                test_run_id=target_test_run_id, actor=actor)
            rc = _cas("bound_to_test_run",
                      target_test_run_id=target_test_run_id,
                      revision_batch_id=rev, reason=reason)
            if rc == 0:
                _conflict()
            _record("bind_test_run",
                    {"target_test_run_id": target_test_run_id,
                     "revision_batch_id": rev, "reason": reason})
        elif action == "request_release":
            if state == "release_requested" and \
                    cur["requested_by"] == actor:
                return self.adjudication_dto(batch_id)
            if state not in ("quarantined", "retained_for_evidence"):
                _invalid()
            rc = _cas("release_requested", requested_by=actor,
                      requested_at=now, reason=reason)
            if rc == 0:
                _conflict()
            _record("request_release", {"reason": reason})
        elif action == "approve_release":
            if state != "release_requested":
                _invalid()
            if actor == cur["requested_by"]:
                raise AdjudicationError(
                    "ADJUDICATION_SAME_ACTOR: 审批人不得与申请人相同"
                    f"（{actor}）")
            rev = self._create_release_revision(
                b, data_scope="operational", test_run_id="",
                actor=actor)
            rc = _cas("release_approved", approved_by=actor,
                      approved_at=now, revision_batch_id=rev,
                      reason=reason)
            if rc == 0:
                _conflict()
            _record("approve_release",
                    {"revision_batch_id": rev, "reason": reason})
        elif action == "reject_release":
            if state != "release_requested":
                _invalid()
            rc = _cas("quarantined", reason=reason)
            if rc == 0:
                _conflict()
            _record("reject_release", {"reason": reason})
        else:
            raise AdjudicationError(
                f"ADJUDICATION_UNKNOWN_ACTION: {action}")
        return self.adjudication_dto(batch_id)
